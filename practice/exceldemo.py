import openpyxl
book = openpyxl.load_workbook("C:\\Users\\rajiv\\OneDrive\\Desktop\\Book1.xlsx")
sheet = book.active
cell = sheet.cell(row=1,column=2)
print(cell.value)

sheet.cell(row=3,column=2).value = "monam"
print(sheet.cell(row=3,column=2).value)
print(sheet.max_row)
print(sheet.max_column)
print(sheet['A7'].value)
for i in range(1,sheet.max_row+1):
    if sheet.cell(row=1 ,column=2).value =="name":

        for j in range(1,sheet.max_column+1):
         print(sheet.cell(row=i,column=j).value)